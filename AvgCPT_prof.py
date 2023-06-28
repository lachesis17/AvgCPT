from PyQt5 import QtWidgets, uic, QtCore, QtGui
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
import time
import math
import pyqtgraph as pg
import sys
import os
import pyperclip
import pandas as pd
import pyodbc
import statistics
import numpy as np
import configparser
import openpyxl
from common.designprofile_wip import DesignProfile
from scipy import stats
from matplotlib import pyplot as plt
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")
QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough) 
QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()

        uic.loadUi("assets/ui/mainwindow.ui", self)
        self.setWindowIcon(QtGui.QIcon('assets/images/icon.ico'))
    
        self.player = QMediaPlayer()
        self.config = configparser.ConfigParser()
        self.config.read('assets/settings.ini')
        self.dark_mode_button.setChecked(bool(self.config.get('Theme','dark')))

        self.cpt_value = ""
        self.pdf_location = ""

        self.button_copy_actual.setIcon(QtGui.QIcon('assets/images/copy.png'))
        self.button_copy_avg.setIcon(QtGui.QIcon('assets/images/copy.png'))
        self.button_copy_actual.setIconSize(QtCore.QSize(25,25))
        self.button_copy_avg.setIconSize(QtCore.QSize(25,25))
        self.actual_val.setReadOnly(True)
        self.average_val.setReadOnly(True)
        self.unit_textbox.setReadOnly(True)
        self.quant_desc.setReadOnly(True)
        self.dir_box.setReadOnly(True)

        self.dark_mode_button.clicked.connect(self.dark_toggle)
        self.full_bh.clicked.connect(self.plot_full_bh)
        self.pdf_box.clicked.connect(self.pdf_dir)
        self.dir_box.mousePressEvent = self.pdf_dir
        self.file_open.triggered.connect(self.get_file_location)
        self.unit_selector.valueChanged.connect(self.change_unit)
        self.button_gint.clicked.connect(self.get_file_location)
        self.button_depth.clicked.connect(self.get_cpt_depths)
        self.button_cpt_val.clicked.connect(self.get_avg_val)
        self.button_copy_actual.clicked.connect(self.copy_actual_value)
        self.button_copy_avg.clicked.connect(self.copy_average_value)
        self.remove_before.clicked.connect(self.remove_data_before)
        self.remove_after.clicked.connect(self.remove_data_after)
        self.remove_at.clicked.connect(self.remove_data_at)
        self.re_plot.clicked.connect(self.recalc_avg)
        self.increment.clicked.connect(self.strain_num_incr)
        self.decrement.clicked.connect(self.strain_num_decr)
        self.full_export.clicked.connect(self.export_full_averages)
        self.actionExport_Averages_for_All.triggered.connect(self.export_full_averages)
        
        #set window size and graph
        self.installEventFilter(self)
        self.set_size()
        self.setup_graph()
        self.dark_mode()



    def setup_graph(self):
        self.plot_area.setContentsMargins(5, 5, 5, 5)
        _labels = {'left': ('Depth', 'm'),'bottom': ('','')}
        self.graph_plot = self.plot_area.addPlot(labels=_labels)
        self.graph_plot.showGrid(x=True, y=True)
        self.graph_plot.setContentsMargins(5,5,5,5)
        if self.dark_mode_button.isChecked() == True:
            self.plot_area.setBackground("#353535")
            self.graph_plot.getAxis('left').setTextPen('white')
            self.graph_plot.getAxis('bottom').setTextPen('white')
        else:
            self.plot_area.setBackground("#f0f0f0")
            self.graph_plot.getAxis('left').setTextPen('black')
            self.graph_plot.getAxis('bottom').setTextPen('black')

    def get_file_location(self):
        #get gint location
        self.gint = ""
        self.cpt_data = ""
        self.reset_graph()
        if not self.config.get('LastFolder','dir') == "":
            self.file_location = QtWidgets.QFileDialog.getOpenFileNames(self,'Open gINT Project', self.config.get('LastFolder','dir'), '*.gpj')
        else:
            self.file_location = QtWidgets.QFileDialog.getOpenFileNames(self,'Open gINT Project', os.getcwd(), '*.gpj')
        try:
            self.file_location = self.file_location[0][0]
            self.point_table.setEnabled(True)
            self.depth_table.setEnabled(True)
            self.full_export.setEnabled(True)
            last_dir = str(os.path.dirname(self.file_location))
            self.config.set('LastFolder','dir',last_dir)
            with open('assets/settings.ini', 'w') as configfile: 
                self.config.write(configfile)
        except:
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Information)
            msgBox.setText("You must select a gINT")
            msgBox.setWindowTitle("No gINT selected")
            msgBox.setStandardButtons(QMessageBox.Ok)
            msgBox.exec()
            self.button_gint.setEnabled(True)
            self.disable_buttons()
            self.point_table.setEnabled(False)
            self.depth_table.setEnabled(False)
            self.full_export.setEnabled(False)
            return
        print(f"Opening {self.file_location}...")

        #check the dir exists
        if os.path.exists(os.path.dirname((self.file_location))):    
            self.gintpath = os.path.dirname((self.file_location))
        else:
            raise ValueError("Please check the directory is correct.")

        #establish connection to sql database (gint)
        try:
            self.gint = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+self.file_location+';')
        except Exception as e:
            print(f"Couldn't establish connection with gINT. Please ensure you have Access Driver 64-bit installed. {e}")
            self.disable_buttons()
            return
        
        print(f"Loaded gINT.")

        bh_query = "SELECT PointID FROM POINT"
        point_list = pd.read_sql(bh_query, self.gint)
        point_id = point_list['PointID'].tolist()
        point_id = sorted(point_id)

        self.point_table.clear()
        self.depth_table.clear()
        for x in point_id:
            item = QListWidgetItem(x)
            item.setTextAlignment(Qt.AlignHCenter)
            self.point_table.addItem(item) 
        self.button_depth.setEnabled(True)
        self.dark_mode()

    def get_cpt_depths(self):
        self.reset_graph()
        self.bh_select = ""
        self.average_val.clear()
        self.geol_layers.clear()
        self.avg_vals.clear()
        self.average_val.clear()
        self.actual_val.clear()
        self.cpt_table.clear()
        self.depth_table.clear()

        try:
            self.bh_select = self.point_table.currentItem().text()
        except AttributeError:
            print("No borehole selected.")
            return

        query = f"SELECT * FROM STCN_DATA WHERE PointID ='{str(self.bh_select)}'"
        self.cpt_data = pd.read_sql(query, self.gint)

        if self.cpt_data.empty:
            print(f'no cpt data for this bh: {self.bh_select}')
            return
        
        self.cpt_data.drop(['GintRecID'], axis=1, inplace=True)
        self.cpt_data.reset_index(inplace=True)
        self.cpt_data.drop(columns=['index'], inplace=True)

        if 'true_depth' not in self.cpt_data:
            self.cpt_data.insert(len(list(self.cpt_data.columns)),'true_depth','')

        self.cpt_data['true_depth'] = self.cpt_data['Depth'] + self.cpt_data['STCN_Depth']
        self.cpt_data['true_depth'] = self.cpt_data['true_depth'].round(2)
        self.cpt_data.sort_values(by=['true_depth'], inplace=True)
        self.cpt_data['true_depth'] = self.cpt_data['true_depth'].map('{:,.2f}'.format)

        depth_float = pd.to_numeric(self.cpt_data['true_depth'])
        self.cpt_data['true_depth'] = depth_float
 
        depth_list = list(self.cpt_data['true_depth'])
        
        self.full_depth = []
        for x in range(0,len(depth_list)):
            self.full_depth.append(round(float(depth_list[x]),2))

        for x in range(0, len(depth_list)):
            depth_list[x] = str(depth_list[x])
            item = QListWidgetItem(depth_list[x])
            item.setTextAlignment(Qt.AlignHCenter)
            self.depth_table.addItem(item) 

        cpt_headers = list(self.cpt_data.columns)
        del cpt_headers[-1], cpt_headers[0], cpt_headers[0], cpt_headers[0], cpt_headers[0],

        self.cpt_table.addItems(cpt_headers)
        self.cpt_table.setCurrentIndex(0)
        self.cpt_table.setEnabled(True)
        self.button_cpt_val.setEnabled(True)
        
        if self.unit_selector.value() == 0:
            self.geol_unit = "GEOL_GEOL"
            self.unit_textbox.clear()
            self.unit_textbox.setText(f'''<p align="center">Geol unit:
{self.geol_unit}</p>''')
        elif self.unit_selector.value() == 1:
            self.geol_unit = "GEOL_GEO2"
            self.unit_textbox.clear()
            self.unit_textbox.setText(f'''<p align="center">Geol unit:
{self.geol_unit}</p>''')
            

    def change_unit(self):
        if self.unit_selector.value() == 0:
            self.geol_unit = "GEOL_GEOL"
            self.unit_textbox.clear()
            self.unit_textbox.setText(f'''<p align="center">Geol unit:
{self.geol_unit}</p>''')
        elif self.unit_selector.value() == 1:
            self.geol_unit = "GEOL_GEO2"
            self.unit_textbox.clear()
            self.unit_textbox.setText(f'''<p align="center">Geol unit:
{self.geol_unit}</p>''')

    def get_geol_layers(self, bh, depth):
        self.bh = bh
        query = f"SELECT * FROM GEOL WHERE PointID ='{str(bh)}'"
        self.geol = pd.read_sql(query, self.gint)

        if self.geol.empty:
            print(f'no geol for this bh: {bh}')
            return

        self.geol.sort_values(by=['Depth'], inplace=True)
        self.geol.drop(['GintRecID'], axis=1, inplace=True)
        self.geol.reset_index(inplace=True)
        self.geol.drop(columns=['index'], inplace=True)

        layer_top = list(self.geol['Depth'])
        layer_base = list(self.geol['GEOL_BASE'])
        layer_leg = list(self.geol['GEOL_LEG'])

        for x in range(0, len(layer_leg)):
            if not str(layer_leg[x]) == "" and "-" in str(layer_leg[x]):
                layer_leg[x] = str(layer_leg[x]).split("-")[1]
            else:
                pass

        layers = list(zip(layer_top, layer_base))
        units = list(self.geol[self.geol_unit])

        if not 'units' in locals():
            units = list(self.geol[self.geol_unit])

        soil_units = list(zip(units, layer_leg))
            
        if not all(v for v in units):
            print("No unitisation - using Geology Legend")
            self.unitised_layers = dict(zip(layers,soil_units))
        else:
            self.unitised_layers = dict(zip(layers,soil_units))

        self.geol_layers_list = []

        if depth == None:
            for (layer, unit) in self.unitised_layers.items():
                qc_list = None
                fs_list = None
                u_list = None
                qnet_list = None
                fr_list = None
                ic_list = None

                # BQ = (STCN_UCOR /1000) / STCN_Qnet

                depth = round((float(layer[0]) + float(layer[1])) / 2,2)
                qc_list = self.cpt_data.loc[(self.cpt_data['true_depth'] >= float(layer[0])) & (self.cpt_data['true_depth'] <= float(layer[1])), ['true_depth','STCN_QC']]
                fs_list = self.cpt_data.loc[(self.cpt_data['true_depth'] >= float(layer[0])) & (self.cpt_data['true_depth'] <= float(layer[1])), ['true_depth', 'STCN_FS']]
                u_list = self.cpt_data.loc[(self.cpt_data['true_depth'] >= float(layer[0])) & (self.cpt_data['true_depth'] <= float(layer[1])), ['true_depth', 'STCN_U']]
                qnet_list = self.cpt_data.loc[(self.cpt_data['true_depth'] >= float(layer[0])) & (self.cpt_data['true_depth'] <= float(layer[1])), ['true_depth', 'STCN_Qnet']]
                fr_list = self.cpt_data.loc[(self.cpt_data['true_depth'] >= float(layer[0])) & (self.cpt_data['true_depth'] <= float(layer[1])), ['true_depth', 'STCN_FCRO']]
                ic_list = self.cpt_data.loc[(self.cpt_data['true_depth'] >= float(layer[0])) & (self.cpt_data['true_depth'] <= float(layer[1])), ['true_depth', 'STCN_SBTi']]

                self.fs_dict[f'Layers'] = ['fs profile']
                self.u_dict[f'Layers'] = ['u profile']
                self.qnet_dict[f'Layers'] = ['qnet profile']
                self.fr_dict[f'Layers'] = ['fr profile']
                self.ic_dict[f'Layers'] = ['ic profile']

                qc_profile = DesignProfile.profile(
                    param=qc_list['STCN_QC'], 
                    depth=qc_list['true_depth'], 
                    name = f'qc (kPa) — {bh} {layer[0]}m to {layer[1]}m - {unit[1]} {unit[0]}', 
                    model=self.get_model(), 
                    zvalue=self.quant_box.value(), 
                    plot=self.pdf_box.isChecked(),
                    save=self.pdf_location)

                fs_profile = DesignProfile.profile(
                    param=fs_list['STCN_FS'], 
                    depth=fs_list['true_depth'], 
                    name = f'fs (MPa) — {bh} {layer[0]}m to {layer[1]}m - {unit[1]} {unit[0]}',
                    model=self.get_model(), 
                    zvalue=self.quant_box.value(), 
                    plot=self.pdf_box.isChecked(),
                    save=self.pdf_location)

                u_profile = DesignProfile.profile(param=u_list['STCN_U'], 
                    depth=qc_list['true_depth'], 
                    name = f'u (kPa) — {bh} {layer[0]}m to {layer[1]}m - {unit[1]} {unit[0]}',
                    model=self.get_model(), 
                    zvalue=self.quant_box.value(), 
                    plot=self.pdf_box.isChecked(),
                    save=self.pdf_location)

                qnet_profile = DesignProfile.profile(param=qnet_list['STCN_Qnet'], 
                    depth=qnet_list['true_depth'], 
                    name = f'qnet (MPa) — {bh} {layer[0]}m to {layer[1]}m - {unit[1]} {unit[0]}',
                    model=self.get_model(),
                    zvalue=self.quant_box.value(), 
                    plot=self.pdf_box.isChecked(),
                    save=self.pdf_location)

                fr_profile = DesignProfile.profile(param=fr_list['STCN_FCRO'], 
                    depth=fr_list['true_depth'], 
                    name = f'fr (-) — {bh} {layer[0]}m to {layer[1]}m - {unit[1]} {unit[0]}',
                    model=self.get_model(), 
                    zvalue=self.quant_box.value(), 
                    plot=self.pdf_box.isChecked(),
                    save=self.pdf_location)

                ic_profile = DesignProfile.profile(param=ic_list['STCN_SBTi'], 
                    depth=ic_list['true_depth'], 
                    name = f'ic (-) — {bh} {layer[0]}m to {layer[1]}m - {unit[1]} {unit[0]}',
                    model=self.get_model(), 
                    zvalue=self.quant_box.value(), 
                    plot=self.pdf_box.isChecked(),
                    save=self.pdf_location)

                self.qc_dict[f'{bh}|{layer[0]}m to {layer[1]}m - {unit[1]}'] = [F"{qc_profile}"]#, qc_list['STCN_QC'].mean(),qc_list['STCN_QC'].std()]
                self.fs_dict[f'{bh}|{layer[0]}m to {layer[1]}m - {unit[1]}'] = [F"{fs_profile}"]#, fs_list['STCN_FS'].mean(),fs_list['STCN_FS'].std()]
                self.u_dict[f'{bh}|{layer[0]}m to {layer[1]}m - {unit[1]}'] = [F"{u_profile}"]#,u_list['STCN_U'].mean(),u_list['STCN_U'].std()]
                self.qnet_dict[f'{bh}|{layer[0]}m to {layer[1]}m - {unit[1]}'] = [f'{qnet_profile}']#qnet_list['STCN_Qnet'].mean(),qnet_list['STCN_Qnet'].std()]
                self.fr_dict[f'{bh}|{layer[0]}m to {layer[1]}m - {unit[1]}'] = [f'{fr_profile}']#fr_list['STCN_FCRO'].mean(),fr_list['STCN_FCRO'].std()]
                self.ic_dict[f'{bh}|{layer[0]}m to {layer[1]}m - {unit[1]}'] = [f'{ic_profile}']#ic_list['STCN_SBTi'].mean(),ic_list['STCN_SBTi'].std()]

                if float(depth) >= float(layer[0]) and float(depth) <= float(layer[1]):
                    if unit[0] == "":
                        self.layer = layer
                        self.unit = "Un-unitised"
                    else:
                        self.layer = layer
                        self.unit = unit[0]
                split_layer_top = str(layer).split(",")[0]
                split_layer_bot = str(layer).split(",")[1]
                split_layer_bot = split_layer_bot[:-1]
                if unit[0] == "":
                    self.geol_layers_list.append(str(f"{split_layer_top}m - {split_layer_bot}m) Soil Type: {unit[1]}"))
                else:
                    self.geol_layers_list.append(str(f"{split_layer_top}m - {split_layer_bot}m) Unit: {unit[0]} | Soil Type: {unit[1]}"))

            return print(f"""~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
LAYERS: {self.geol_layers_list}
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~""")

        for (layer, unit) in self.unitised_layers.items():
            if float(depth) >= float(layer[0]) and float(depth) <= float(layer[1]):
                if unit[0] == "":
                    print(f"The depth {depth}m is in a {unit[1]} layer {layer} and the unitisation has not been done...")
                    self.layer = layer
                    self.unit = "Un-unitised"
                else:
                    print(f"The depth {depth}m is in a {unit[1]} layer {layer} and the unit is {unit[0]}.")
                    self.layer = layer
                    self.unit = unit[0]
            split_layer_top = str(layer).split(",")[0]
            split_layer_bot = str(layer).split(",")[1]
            split_layer_bot = split_layer_bot[:-1]
            if unit[0] == "":
                self.geol_layers_list.append(str(f"{split_layer_top}m - {split_layer_bot}m) Soil Type: {unit[1]}"))
            else:
                self.geol_layers_list.append(str(f"{split_layer_top}m - {split_layer_bot}m) Unit: {unit[0]} | Soil Type: {unit[1]}"))

        print(f"""~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
LAYERS: {self.geol_layers_list}
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~""")
            
        self.geol_layers.clear()
        self.geol_layers.addItems(self.geol_layers_list)
        self.geol_layers.setEnabled(True)


    def export_full_averages(self):
        self.qc_dict = {}
        self.fs_dict = {}
        self.u_dict = {}
        self.qnet_dict = {}
        self.fr_dict = {}
        self.ic_dict = {}
        self.change_unit()
    
        bhs_in_gint = [self.point_table.item(x).text() for x in range(self.point_table.count())]

        if (self.pdf_box.isChecked() == True and self.pdf_location == "") or (self.pdf_box.isChecked() == True and self.dir_box.text() == "Please select a PDF save directory!"):
            self.dir_box.setText(f'Please select a PDF save directory!')
            print("Please select a directory for PDF export.")
            return

        for x in range(0,len(bhs_in_gint)):

            query = f"SELECT * FROM STCN_DATA WHERE PointID ='{str(bhs_in_gint[x])}'"
            self.cpt_data = pd.read_sql(query, self.gint)

            if self.cpt_data.empty:
                print(f'no cpt data for this bh: {bhs_in_gint[x]}')
                pass
            
            #add true depth
            self.cpt_data.drop(['GintRecID'], axis=1, inplace=True)
            self.cpt_data.reset_index(inplace=True)
            self.cpt_data.drop(columns=['index'], inplace=True)
            self.cpt_data.insert(len(list(self.cpt_data.columns)),'true_depth','')
            self.cpt_data['true_depth'] = self.cpt_data['Depth'] + self.cpt_data['STCN_Depth']
            self.cpt_data['true_depth'] = self.cpt_data['true_depth'].round(2)
            self.cpt_data.sort_values(by=['true_depth'], inplace=True)
            self.cpt_data['true_depth'] = self.cpt_data['true_depth'].map('{:,.2f}'.format)
            depth_float = pd.to_numeric(self.cpt_data['true_depth'])
            self.cpt_data['true_depth'] = depth_float

            #loop through each bh and add vals to dict
            self.get_geol_layers(bh=bhs_in_gint[x], depth=None)
            QApplication.processEvents()

        #build dict with keys as index - needs to use these as index for 'scalar array' error
        self.full_df = pd.DataFrame.from_dict(self.qc_dict, orient='index', columns=['qc profile'])
        self.full_df['Borehole'] = self.full_df.index
        self.full_df[['Borehole','Geology Layers']] = self.full_df['Borehole'].str.split('|', expand=True)
        move_cols = ['Borehole','Geology Layers']
        self.full_df = self.full_df[move_cols + [col for col in self.full_df.columns if col not in move_cols]]

        #add in the rest of the value dictionaries into full df
        def build_df_from_dict(y):
            _data = [v for k,v in y.items()]
            df = pd.DataFrame.from_dict(y, orient='index', columns=_data[0])   
            df = df.iloc[1:]         
            col_list = [x for x in df.columns]
            self.full_df[col_list[0]] = df[col_list[0]]
            #self.full_df[col_list[1]] = df[col_list[1]]

        build_df_from_dict(self.fs_dict)
        build_df_from_dict(self.u_dict)
        build_df_from_dict(self.qnet_dict)
        build_df_from_dict(self.fr_dict)
        build_df_from_dict(self.ic_dict)

        keep_cols = [col for col in self.full_df.columns if 'std' not in col]
        self.full_df = self.full_df[keep_cols + [col for col in self.full_df.columns if 'std' in col]]

        self.full_df.replace(to_replace=0, value=np.nan, inplace=True)
        self.full_df.replace(to_replace="None", value=np.nan, inplace=True) 
        self.full_df.replace({'\\[': ''}, regex=True, inplace=True)
        self.full_df.replace({'\\]': ''}, regex=True, inplace=True)
        self.full_df.dropna(axis = 1, how="all", inplace= True)

        self.full_df[["Qc Best Estimate TOP",
                      "Qc Best Estimate BOT",
                      "Qc Lower Bounds TOP",
                      "Qc Lower Bounds BOT",
                      "Qc Upper Bounds TOP",
                      "Qc Upper Bounds BOT",
                      "Qc Mean at 95% Confidence (Upper) - TOP",
                      "Qc Mean at 95% Confidence (Upper) - BOT",
                      "Qc Mean at 95% Confidence (Lower) - TOP",
                      "Qc Mean at 95% Confidence (Lower) - BOT",
                      "Qc Standard Deviation - All Data",
                      "Qc Standard Deviation - Outliers Removed"]] = self.full_df['qc profile'].str.split(',', expand=True)
        self.full_df[["Fs Best Estimate TOP",
                      "Fs Best Estimate BOT",
                      "Fs Lower Bounds TOP",
                      "Fs Lower Bounds BOT",
                      "Fs Upper Bounds TOP",
                      "Fs Upper Bounds BOT",
                      "Fs Mean at 95% Confidence (Upper) - TOP",
                      "Fs Mean at 95% Confidence (Upper) - BOT",
                      "Fs Mean at 95% Confidence (Lower) - TOP",
                      "Fs Mean at 95% Confidence (Lower) - BOT",
                      "Fs Standard Deviation - All Data",
                      "Fs Standard Deviation - Outliers Removed"]] = self.full_df['fs profile'].str.split(',', expand=True)
        self.full_df[["U Best Estimate TOP",
                      "U Best Estimate BOT",
                      "U Lower Bounds TOP",
                      "U Lower Bounds BOT",
                      "U Upper Bounds TOP",
                      "U Upper Bounds BOT",
                      "U Mean at 95% Confidence (Upper) - TOP",
                      "U Mean at 95% Confidence (Upper) - BOT",
                      "U Mean at 95% Confidence (Lower) - TOP",
                      "U Mean at 95% Confidence (Lower) - BOT",
                      "U Standard Deviation - All Data",
                      "U Standard Deviation - Outliers Removed"]] = self.full_df['u profile'].str.split(',', expand=True)  
        self.full_df[["Qnet Best Estimate TOP",
                      "Qnet Best Estimate BOT",
                      "Qnet Lower Bounds TOP",
                      "Qnet Lower Bounds BOT",
                      "Qnet Upper Bounds TOP",
                      "Qnet Upper Bounds BOT",
                      "Qnet Mean at 95% Confidence (Upper) - TOP",
                      "Qnet Mean at 95% Confidence (Upper) - BOT",
                      "Qnet Mean at 95% Confidence (Lower) - TOP",
                      "Qnet Mean at 95% Confidence (Lower) - BOT",
                      "Qnet Standard Deviation - All Data",
                      "Qnet Standard Deviation - Outliers Removed"]] = self.full_df['qnet profile'].str.split(',', expand=True)
        self.full_df[["Fr Best Estimate TOP",
                      "Fr Best Estimate BOT",
                      "Fr Lower Bounds TOP",
                      "Fr Lower Bounds BOT",
                      "Fr Upper Bounds TOP",
                      "Fr Upper Bounds BOT",
                      "Fr Mean at 95% Confidence (Upper) - TOP",
                      "Fr Mean at 95% Confidence (Upper) - BOT",
                      "Fr Mean at 95% Confidence (Lower) - TOP",
                      "Fr Mean at 95% Confidence (Lower) - BOT",
                      "Fr Standard Deviation - All Data",
                      "Fr Standard Deviation - Outliers Removed"]] = self.full_df['fr profile'].str.split(',', expand=True) 
        if 'ic profile' in self.full_df.columns:
            self.full_df[["Ic Best Estimate TOP",
                        "Ic Best Estimate BOT",
                        "Ic Lower Bounds TOP",
                        "Ic Lower Bounds BOT",
                        "Ic Upper Bounds TOP",
                        "Ic Upper Bounds BOT",
                        "Ic Mean at 95% Confidence (Upper) - TOP",
                        "Ic Mean at 95% Confidence (Upper) - BOT",
                        "Ic Mean at 95% Confidence (Lower) - TOP",
                        "Ic Mean at 95% Confidence (Lower) - BOT",
                        "Ic Standard Deviation - All Data",
                        "Ic Standard Deviation - Outliers Removed"]] = self.full_df['ic profile'].str.split(',', expand=True)  
            self.full_df.drop(columns=['qc profile', 'fs profile', 'u profile', 'qnet profile', 'fr profile', 'ic profile'], inplace=True)
        else:
            self.full_df.drop(columns=['qc profile', 'fs profile', 'u profile', 'qnet profile', 'fr profile'], inplace=True)

        fname = QtWidgets.QFileDialog.getSaveFileName(self, "Save export of CPT averages...", os.getcwd(), "Excel file *.xlsx;; CSV *.csv")
        
        if fname[0] == '':
            return
        
        if fname[1] == 'Excel file *.xlsx':
            self.full_df.to_excel(fname[0], sheet_name="Average CPT Values", index=False)

            wb = openpyxl.load_workbook(f"{fname[0]}")
            ws = wb.active
            ws.insert_rows(1)

            for column_cells in ws.columns:
                column_length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = (get_column_letter(column_cells[0].column))
                last_col = column_letter
                ws.column_dimensions[column_letter].width = column_length * 1.15
            #ws.column_dimensions['A'].width = column_length * 1.25
            
            ws.merge_cells('A1:A2')
            ws['A1'] = 'Borehole'
            ws['A1'].font = Font(bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('B1:B2')
            ws['B1'] = 'Geology Layers'
            ws['B1'].font = Font(bold=True)
            ws['B1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('C1:N1')
            ws['C1'] = 'Qc CPT Profile'
            ws['C1'].font = Font(bold=True) 
            ws.merge_cells('O1:Z1')
            ws['O1'] = 'Fs CPT Profile'
            ws['O1'].font = Font(bold=True) 
            ws.merge_cells('AA1:AL1')
            ws['AA1'] = 'U CPT Profile'
            ws['AA1'].font = Font(bold=True) 
            ws.merge_cells('AM1:AX1')
            ws['AM1'] = 'Qnet CPT Profile'
            ws['AM1'].font = Font(bold=True) 
            ws.merge_cells('AY1:BJ1')
            ws['AY1'] = 'Fr CPT Profile'
            ws['AY1'].font = Font(bold=True) 
            ws.merge_cells('BK1:BV1')
            ws['BK1'] = 'Ic CPT Profile'
            ws['BK1'].font = Font(bold=True) 

            def set_border(ws, cell_range):
                thin = Side(border_style="thin", color="000000")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            def set_single_border(ws, cell_range):
                thin = Side(border_style="thin", color="000000")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = Border(top=None, left=thin, right=None, bottom=None)

            end_row = str(len(ws['A']))
            border_range = 'C1:C' + end_row
            set_single_border(ws, border_range) 
            border_range = 'O1:O' + end_row
            set_single_border(ws, border_range) 
            border_range = 'AA1:AA' + end_row
            set_single_border(ws, border_range) 
            border_range = 'AM1:AM' + end_row
            set_single_border(ws, border_range) 
            border_range = 'AY1:AY' + end_row
            set_single_border(ws, border_range) 
            border_range = 'BK1:BK' + end_row
            set_single_border(ws, border_range) 
            border_range = 'BW1:BW' + end_row
            set_single_border(ws, border_range) 

            set_border(ws, 'A1:BV2')
            head_row = ws['C3']
            ws.freeze_panes = head_row

            wb.save(f"{fname[0]}")

        elif fname[1] == 'CSV *.csv':
            self.full_df.to_csv(fname[0], index=False)

        print(f"All borehole's average CPT values saved in: {fname[0]}")

        return

    def get_avg_val(self):
        self.full_bh.setChecked(False)
        self.cpt_value = self.cpt_table.currentText()
        rows = self.cpt_data.shape[0]
        avg_list = []
        depth_with_value = {}
        avg_row: int
        zero_count = 0
        max_count = 0
        min_count = 0

        try:
            self.cpt_depth = self.depth_table.currentItem().text()
        except AttributeError:
            print("No depth selected.")
            return

        print(f"""****************************************           
Querying {self.cpt_value} at {self.cpt_depth}m from {self.bh_select}""")

        self.get_geol_layers(bh=self.bh_select, depth=self.cpt_depth)

        min_datapoint = 0
        if not rows - 1 == -1:
            max_datapoint = rows - 1
        for row in range(0, rows):
            #these nan checks mean that if there is no data in -0.5m (e.g. BHCPT67A 18.90m) it can't define min_datapoint, so its kept as 0 and goes to bottom of hole
            if self.cpt_data['true_depth'][row] == None or self.cpt_data['true_depth'][row] == "":
                pass
            elif np.isnan(float(self.cpt_data['true_depth'][row])) == True:
                pass
            else:
                if float(self.cpt_data['true_depth'][row]) <= (float(self.cpt_depth) - float(0.5)):
                    min_datapoint = row
                elif float(self.cpt_data['true_depth'][row]) >= (float(self.cpt_depth) + float(0.5)) and not float(self.cpt_data['true_depth'][row]) > (float(self.cpt_depth) + float(0.51)):
                    max_datapoint = row
                elif float(self.cpt_data['true_depth'][row]) > (float(self.cpt_depth) + float(0.51)):
                    break
                #need to somehow implement this so it doesnt take the spikes without not plotting anything if the above depths contain no data
                elif not row == rows -1:
                    if not float(self.cpt_data['true_depth'][row]) < 0.15:
                        if float(self.cpt_data['true_depth'][row+1]) % float(self.cpt_data['true_depth'][row]) > 0.15:
                            if float(self.cpt_data['true_depth'][row]) == float(self.cpt_depth):
                                self.reset_graph()
                                return print("this depth is in a data gap")
                            #need to understand why it cant get avg_row - 22.88m BHCPT67A
                            min_datapoint += 1 
                            max_datapoint = row
                            break

            if str(self.cpt_data['true_depth'][row]) == str(self.cpt_depth):
                avg_row = row         
                cpt_result = self.cpt_data[self.cpt_value][row]
                self.actual_val.clear()
                self.actual_val.setText(f'''<p align="center">{self.bh_select} value for {self.cpt_value} at {self.cpt_depth}m is: {self.cpt_data[self.cpt_value][row]}</p>''')
                print(f"The value for {self.cpt_value} at {self.cpt_depth}m is: {self.cpt_data[self.cpt_value][row]}")

        # print(f"min row: {min_datapoint} - data is: {self.cpt_data[self.cpt_value][min_datapoint]} - depth is: {self.cpt_data['true_depth'][min_datapoint]}")
        # print(f"avg row: {avg_row} - data is: {self.cpt_data[self.cpt_value][avg_row]} - depth is: {self.cpt_data['true_depth'][avg_row]}")
        # print(f"max row: {max_datapoint} - data is: {self.cpt_data[self.cpt_value][max_datapoint]} - depth is: {self.cpt_data['true_depth'][max_datapoint]}")


        #can be used in next loops when iterating to check if true_depth[row] is in layer
        for layer in range(0, len(self.geol_layers_list)):
            x = str(self.geol_layers_list[layer]).split("-")[0].split("(")[1].split("m")[0]
            if x == str(self.layer[0]):
                self.geol_layers.setCurrentIndex(layer) 

        for row in range(min_datapoint, max_datapoint):
            if float(self.cpt_data['true_depth'][row]) < (float(self.cpt_depth) - float(0.51)):
                min_datapoint += 1

        for row in range(min_datapoint, max_datapoint):
            if row <= 0:
                zero_count += 1
                pass
            elif row >= rows:
                max_count = row - rows
                pass
            #this line means that if there is no data in +0.5m (e.g. BHCPT67A 18.90m) it can't append max_count as value == None (resolved)
            elif not self.cpt_data[self.cpt_value][row] == None and not self.cpt_data[self.cpt_value][row] == "":
                if float(self.cpt_data['true_depth'][row]) >= float(self.layer[1]) and not float(self.cpt_data['true_depth'][row]) > (float(self.cpt_depth) + float(0.51)):
                    print(f"""Average data ranges for {self.cpt_data['true_depth'][row]} exceeds layer range of {self.layer}, cutting data...""")
                    max_count +=1
                    pass
                elif float(self.cpt_data['true_depth'][row]) < float(self.layer[0]) and not float(self.cpt_data['true_depth'][row]) < (float(self.cpt_depth) - float(0.51)):
                    print(f"""Average data ranges for {self.cpt_data['true_depth'][row]} precedes layer range of {self.layer}, cutting data...""")
                    min_count += 1
                    pass
                # elif float(self.cpt_data['true_depth'][row]) % float(self.cpt_data['true_depth'][avg_row]) < 0.5:
                #     min_count += 1
                #does this work? - nope
                else:
                    avg_list.append(float(self.cpt_data[self.cpt_value][row]))
                    depth_with_value[f"{self.cpt_data['true_depth'][row]}m"] = self.cpt_data[self.cpt_value][row]

        #check if layer is less than 1m otherwise the +/- 0.5m range cannot get correct index of vals
        if self.layer[1] - self.layer[0] < float(0.999):
            zero_count = 0
            min_count = 0
            max_count = 0

        #if -0.5m puts range before zero
        if not zero_count == 0:
            for row in range (0 + zero_count, (max_datapoint) + zero_count):
                if row >= rows:
                    break
                elif not self.cpt_data[self.cpt_value][row] == None and not self.cpt_data[self.cpt_value][row] == "":
                    avg_list.append(float(self.cpt_data[self.cpt_value][row]))
                    depth_with_value[f"{self.cpt_data['true_depth'][row]}m"] = self.cpt_data[self.cpt_value][row]

        #if -0.5m gets cut because of end of layer, append values from above. min_count = the difference of data points exceeding range + (requested depth + 0.5m)
        if not min_count == 0:
            for row in range (min_datapoint + min_count, (max_datapoint) + min_count):
                if row >= rows or row < 0:
                    break
                elif not self.cpt_data[self.cpt_value][row] == None and not self.cpt_data[self.cpt_value][row] == "":
                    avg_list.append(float(self.cpt_data[self.cpt_value][row]))
                    depth_with_value[f"{self.cpt_data['true_depth'][row]}m"] = self.cpt_data[self.cpt_value][row]

        #if -0.5m gets cut because of end of layer or hole, append values from above. max_count = the difference of data points exceeding range - (requested depth - 0.5m)
        if not max_count == 0:
            for row in range(min_datapoint,(min_datapoint - max_count),-1):
                if row >= rows:
                    pass
                elif not self.cpt_data[self.cpt_value][row] == None:
                    avg_list.insert(0, float(self.cpt_data[self.cpt_value][row]))
                    depth_with_value[f"{self.cpt_data['true_depth'][row]}m"] = self.cpt_data[self.cpt_value][row]

        depth_with_value = dict(sorted(depth_with_value.items(), key=lambda x: float(str(x[0]).split("m")[0])))
        avg_list = [x for x in avg_list if not np.isnan(x) and not x == None]

        depth_with_value_str = [str(x).split("('")[1].replace("',", " -").replace(")", "") for x in depth_with_value.items()]

        x_coord = []
        y_coord = []

        for (k,v) in depth_with_value.items():
            y_coord.append(float(str(k).split("m")[0]))
            x_coord.append(float(str(v)))

        for y in range(0, len(y_coord)):
            if float(y_coord[y]) == float(self.cpt_depth):
                self.avg_line = y

        self.avg_vals.clear()
        self.avg_vals.addItems(depth_with_value_str)
        self.button_copy_actual.setEnabled(True)
        self.button_copy_avg.setEnabled(True)

        self.x = x_coord
        self.y = y_coord

        print(f"Depths in a 1(m) range: {y_coord}")
        print(f"Values in a 1(m) range: {x_coord}")

        if avg_list == []:
            print(f"No data - check the data (e.g, is there data? Fs has no data at the end of the push.")
            self.actual_val.clear()
            self.actual_val.setText(f'''<p align="center">No data - check the data (e.g, is there data? Fs has no data at the end of the push.</p>''')
            return
        else:
            for depth, value in depth_with_value.items():
                if str(depth).split("m")[0] == str(self.cpt_depth):
                    self.avg_vals.setCurrentText(str(f"{depth} - {value}"))
                    break
                else:
                    self.avg_vals.setCurrentIndex(0)

        avg_val = statistics.mean(avg_list)
        self.avg_vals.setEnabled(True)

        self.average_val.clear()
        self.average_val.setText(f'''<p align="center">{self.bh_select} average value for {self.cpt_value} at {self.cpt_depth}m is: {round(avg_val, 4)}</p>''')
        print(f"""The average value for {self.cpt_value} at {self.cpt_depth}m is: {avg_val}
****************************************""")
        self.reset_graph()
        self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
        self.play_coin()
        
    def play_coin(self):
        coin_num = np.random.randint(177)
        if coin_num == 17:
            coin = ('assets/sounds/coin.mp3')
            coin_url = QUrl.fromLocalFile(coin)
            content = QMediaContent(coin_url)
            self.player.setMedia(content)
            self.player.setVolume(33)
            self.player.play()

    def play_nice(self):
        nice_num = np.random.randint(77)
        if nice_num == 7:
            nice = ('assets/sounds/nice.mp3')
            nice_url = QUrl.fromLocalFile(nice)
            content = QMediaContent(nice_url)
            self.player.setMedia(content)
            self.player.setVolume(33)
            self.player.play()

    def plot_full_bh(self):
        self.play_coin()
        self.reset_graph()
        self.dark_mode()

        if self.full_bh.isChecked() == True:
            #need to fix all the empty strings (e.g. beacon gint) otherwise .plot() cannot get min and max bounding values for axis as not real number (convert empty string to nan)
            x_fixed = list(self.cpt_data[self.cpt_value])
            for x in range(0, len(x_fixed)):
                if not x_fixed[x] == None:
                    if x_fixed[x] == '' or x_fixed[x] == "":
                        x_fixed[x] = float('nan')
                    else:
                        x_fixed[x] = float(x_fixed[x])
            self.plot_graph(x_fixed, self.full_depth, cpt_value=self.cpt_value)
        else:
            self.plot_graph(self.x, self.y, cpt_value=self.cpt_value)
    
    def plot_graph(self, x, y, cpt_value):
        if self.dark_mode_button.isChecked() == True and self.full_bh.isChecked() == False:
            self.graph_plot.plot(x,y, symbol='o', symbolSize='5', pen='w', symbolPen='r', symbolBrush='r', axisx='w', axisy='w')
        elif self.full_bh.isChecked() == True:
            self.graph_plot.plot(x,y, pen='r',  axisx='w', axisy='w')
        else:
            self.plot_area.setBackground("#f0f0f0")
            self.graph_plot.plot(x,y, symbol='o', symbolSize='5', pen='b', symbolPen='b', symbolBrush='b', axisx='b', axisy='b')
        self.graph_plot.getAxis('bottom').setLabel(f"{cpt_value}")
        self.graph_plot.getViewBox().invertY(True)

    def reset_graph(self):
        self.plot_area.clear()
        self.setup_graph()

    def copy_actual_value(self):
        print("Copied value at depth to clipboard.")
        pyperclip.copy(self.actual_val.toPlainText())

    def copy_average_value(self):
        print("Copied average value to clipboard.")
        pyperclip.copy(self.average_val.toPlainText())

    def remove_data_before(self):
        if hasattr(self, 'line'):
            if self.avg_line == 0:
                del self.y[self.avg_line]
                del self.x[self.avg_line]
                delattr(self, 'line')
            else:
                if abs(0-self.avg_line) > 5:
                    confirm = QMessageBox
                    ask = confirm.question(self, '', f'''You are about to remove a lot of data.
{abs(0-self.avg_line)} data points in total.
Are you sure you want to delete this much data?''')
                    if ask == confirm.Yes:
                        del self.y[0:self.avg_line]
                        del self.x[0:self.avg_line]
                        self.reset_graph()
                        self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
                        delattr(self, 'line')
                    else:
                        return
                else:
                    del self.y[0:self.avg_line]
                    del self.x[0:self.avg_line]
                self.reset_graph()
                self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
        else:
            return
        

    def remove_data_after(self):
        if hasattr(self, 'line'):
            if self.avg_line >= len(self.y):
                del self.y[self.avg_line]
                del self.x[self.avg_line]
                delattr(self, 'line')
            else:
                if abs(self.avg_line-len(self.y)) > 5:
                    confirm = QMessageBox
                    ask = confirm.question(self, '', f'''You are about to remove a lot of data.
{abs(self.avg_line-len(self.y))} data points in total.
Are you sure you want to delete this much data?''')
                    if ask == confirm.Yes:
                        del self.y[self.avg_line:len(self.y)]
                        del self.x[self.avg_line:len(self.x)]
                        self.reset_graph()
                        self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
                        delattr(self, 'line')
                    else:
                        return
                else:    
                    del self.y[self.avg_line:len(self.y)]
                    del self.x[self.avg_line:len(self.x)]
            self.reset_graph()
            self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
        else:
            return
        
    def remove_data_at(self):
        if hasattr(self, 'line'):
            del self.y[self.avg_line]
            del self.x[self.avg_line]
            self.reset_graph()
            self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
            delattr(self, 'line')
        else:
            return
        
    def strain_num_incr(self):
        if (self.avg_line + 1) >= len(self.y):
            self.avg_line = 0
        else:
            self.avg_line += 1
        self.add_line()

    def strain_num_decr(self):
        if (self.avg_line - 1) < 0:
            self.avg_line = len(self.y) - 1
        else:
            self.avg_line -= 1
        self.add_line()

    def add_line(self):
        if hasattr(self, 'line'):
            self.graph_plot.removeItem(self.line)
        try:
            self.line = pg.InfiniteLine(pos=self.y[self.avg_line], angle=0, pen=pg.mkPen('r', width=2), movable=False)
            self.graph_plot.addItem(self.line)
        except:
            pass

    def recalc_avg(self):
        x = self.x
        y = self.y
        self.avg_vals.clear()

        new_depth_with_val = []

        for data_point in range(0, len(y)):
            new_depth_with_val.append(f"{y[data_point]}m - {x[data_point]}")

        self.avg_vals.addItems(new_depth_with_val)
        print(new_depth_with_val)
        for item in range(0, len(new_depth_with_val)):
            if str(new_depth_with_val[item]).split("m")[0] == str(self.cpt_depth):
                self.avg_vals.setCurrentText(str(f"{new_depth_with_val[item]}"))
                break
            else:
                self.avg_vals.setCurrentIndex(0)

        avg_val = statistics.mean(self.x)

        self.average_val.clear()
        self.average_val.setText(f'''<p align="center">{self.bh_select} (recalculated) average value for {self.cpt_value} at {self.cpt_depth}m is: {round(avg_val, 4)}</p>''')
        print(f"""Recalculated average value for {self.cpt_value} at {self.cpt_depth}m is: {avg_val}
****************************************""")
        

    def get_model(self):
        model_selection = self.model_box.currentText()

        if model_selection == "Dependent":
            return "DEP"
        if model_selection == "Independent":
            return "IND"
        if model_selection == "Automatic":
            return "AUTO"

        
    def pdf_dir(self,event):
        if self.pdf_box.isChecked() == True:
            self.pdf_location = QtWidgets.QFileDialog.getExistingDirectory(self, "Save PDF of CPT Profiles...", os.getcwd()) 
            if self.pdf_location == "":
                self.dir_box.setText(f'Please select a PDF save directory!')
                return
        self.dir_box.clear()
        self.dir_box.setText(f'{self.pdf_location}')
        if self.pdf_box.isChecked() == False and self.pdf_location == "":
            self.dir_box.setText(f'Please select a PDF save directory!')
            return
        

    def dark_toggle(self):
        self.play_nice()
        self.dark_mode()

    def dark_mode(self):
        if self.dark_mode_button.isChecked() == False:
            #LIGHT THEME
            self.dark_mode_button.setChecked(False)
            self.config.set('Theme','dark','')
            with open('assets/settings.ini', 'w') as configfile: 
                self.config.write(configfile)
            self.plot_area.setBackground("#f0f0f0")
            self.button_copy_actual.setIcon(QtGui.QIcon('assets/images/copy.png'))
            self.button_copy_avg.setIcon(QtGui.QIcon('assets/images/copy.png'))
            
            light_palette = QPalette()
            light_palette.setColor(QPalette.Window, QColor(240, 240, 240))
            light_palette.setColor(QPalette.Background, QColor('#f0f0f0'))
            light_palette.setColor(QPalette.WindowText, Qt.black)
            light_palette.setColor(QPalette.Base, QColor(240, 240, 240))
            light_palette.setColor(QPalette.AlternateBase, QColor(240, 240, 240))
            light_palette.setColor(QPalette.ToolTipBase, QColor(240, 240, 240))
            light_palette.setColor(QPalette.ToolTipText, Qt.black)
            light_palette.setColor(QPalette.Text, Qt.black)#
            light_palette.setColor(QPalette.Button, QColor(240, 240, 240))
            light_palette.setColor(QPalette.ButtonText, Qt.white)#
            light_palette.setColor(QPalette.BrightText, Qt.red)
            light_palette.setColor(QPalette.Link, QColor(42, 130, 218))
            light_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
            light_palette.setColor(QPalette.HighlightedText, QColor(240, 240, 240))
            light_palette.setColor(QPalette.Active, QPalette.Button, QColor(240, 240, 240))
            light_palette.setColor(QPalette.Disabled, QPalette.ButtonText, Qt.lightGray)
            light_palette.setColor(QPalette.Disabled, QPalette.WindowText, Qt.lightGray)
            light_palette.setColor(QPalette.Disabled, QPalette.Text, Qt.lightGray)
            light_palette.setColor(QPalette.Disabled, QPalette.Light, QColor('#f0f0f0'))
            self.tabgroup.setStyleSheet(f"{self.config.get('Theme','tab_css')}")
            self.left_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css_light')}")
            self.top_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css_light')}")
            self.top_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css_light')}")
            self.right_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css_light')}")
            self.bot_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css_light')}")
            self.unit_textbox.setStyleSheet(f"{self.config.get('Theme','textbox_css_light')}")
            self.actual_val.setStyleSheet(f"{self.config.get('Theme','textbox_dark_css_light')}")
            self.average_val.setStyleSheet(f"{self.config.get('Theme','textbox_dark_css_light')}")
            self.button_copy_actual.setStyleSheet(f"{self.config.get('Theme','button_transp_css_light')}")
            self.button_copy_avg.setStyleSheet(f"{self.config.get('Theme','button_transp_css_light')}")
            self.button_gint.setStyleSheet(f"{self.config.get('Theme','button_css_light')}")
            self.button_depth.setStyleSheet(f"{self.config.get('Theme','button_css_light')}")
            self.button_cpt_val.setStyleSheet(f"{self.config.get('Theme','button_css_light')}")
            self.remove_before.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.remove_after.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.remove_at.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.re_plot.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.increment.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.decrement.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.cpt_table.setStyleSheet(f"{self.config.get('Theme','combo_css_light')}")
            self.geol_layers.setStyleSheet(f"{self.config.get('Theme','combo_css_light')}")
            self.avg_vals.setStyleSheet(f"{self.config.get('Theme','combo_css_light')}")
            self.dark_mode_button.setStyleSheet(f"{self.config.get('Theme','checkbox_css_light')}")
            self.full_bh.setStyleSheet(f"{self.config.get('Theme','checkbox_css_light')}")
            self.full_export.setStyleSheet(f"{self.config.get('Theme','button_css_sml_light')}")
            self.menubar.setStyleSheet(f"font: 10pt 'Roboto'; background: #f0f0f0; color: black;")
            self.point_table.setStyleSheet(f"{self.config.get('Theme','table_css_light')}")
            self.depth_table.setStyleSheet(f"{self.config.get('Theme','table_css_light')}")

            self.reset_graph()   
            if not self.cpt_value == "":
                self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
            QApplication.setPalette(light_palette)


        else:
            #DARK THEME
            self.dark_mode_button.setChecked(True)
            self.config.set('Theme','dark','True')
            with open('assets/settings.ini', 'w') as configfile: 
                self.config.write(configfile)
            self.plot_area.setBackground("#353535")
            self.button_copy_actual.setIcon(QtGui.QIcon('assets/images/copy_light.png'))
            self.button_copy_avg.setIcon(QtGui.QIcon('assets/images/copy_light.png'))

            dark_palette = QPalette()
            dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.Background, QColor('#353535'))
            dark_palette.setColor(QPalette.Foreground, QColor('#353535'))
            dark_palette.setColor(QPalette.WindowText, Qt.black)
            dark_palette.setColor(QPalette.Base, QColor(35, 35, 35))
            dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ToolTipBase, QColor(25, 25, 25))
            dark_palette.setColor(QPalette.ToolTipText, Qt.black)
            dark_palette.setColor(QPalette.Text, Qt.white)
            dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ButtonText, Qt.black)
            dark_palette.setColor(QPalette.BrightText, Qt.red)
            dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.Highlight, Qt.darkGray)
            dark_palette.setColor(QPalette.HighlightedText, QColor(35, 35, 35))
            dark_palette.setColor(QPalette.Active, QPalette.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.Disabled, QPalette.ButtonText, Qt.darkGray)
            dark_palette.setColor(QPalette.Disabled, QPalette.WindowText, Qt.darkGray)
            dark_palette.setColor(QPalette.Disabled, QPalette.Text, Qt.darkGray)
            dark_palette.setColor(QPalette.Disabled, QPalette.Light, QColor(53, 53, 53))
            self.tabgroup.setStyleSheet(f"{self.config.get('Theme','tab_css_dark')}")
            self.left_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css')}")
            self.top_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css')}")
            self.top_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css')}")
            self.right_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css')}")
            self.bot_spacer.setStyleSheet(f"{self.config.get('Theme','spacer_css')}")
            self.unit_textbox.setStyleSheet(f"{self.config.get('Theme','textbox_css')}")
            self.actual_val.setStyleSheet(f"{self.config.get('Theme','textbox_dark_css')}")
            self.average_val.setStyleSheet(f"{self.config.get('Theme','textbox_dark_css')}")
            self.button_copy_actual.setStyleSheet(f"{self.config.get('Theme','button_transp_css')}")
            self.button_copy_avg.setStyleSheet(f"{self.config.get('Theme','button_transp_css')}")
            self.button_gint.setStyleSheet(f"{self.config.get('Theme','button_css')}")
            self.button_depth.setStyleSheet(f"{self.config.get('Theme','button_css')}")
            self.button_cpt_val.setStyleSheet(f"{self.config.get('Theme','button_css')}")
            self.remove_before.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.remove_after.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.remove_at.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.re_plot.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.increment.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.decrement.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.cpt_table.setStyleSheet(f"{self.config.get('Theme','combo_css')}")
            self.geol_layers.setStyleSheet(f"{self.config.get('Theme','combo_css')}")
            self.avg_vals.setStyleSheet(f"{self.config.get('Theme','combo_css')}")
            self.dark_mode_button.setStyleSheet(f"{self.config.get('Theme','checkbox_css')}")
            self.full_bh.setStyleSheet(f"{self.config.get('Theme','checkbox_css')}")
            self.full_export.setStyleSheet(f"{self.config.get('Theme','button_css_sml')}")
            self.menubar.setStyleSheet(f"font: 10pt 'Roboto'; background: #353535; color: white;")
            self.point_table.setStyleSheet(f"{self.config.get('Theme','table_css')}")
            self.depth_table.setStyleSheet(f"{self.config.get('Theme','table_css')}")

            self.reset_graph()
            if not self.cpt_value == "":
                self.plot_graph(x=self.x, y=self.y, cpt_value=self.cpt_value)
            QApplication.setPalette(dark_palette)


    def disable_buttons(self):       
        self.button_depth.setEnabled(False)
        self.cpt_table.setEnabled(False)
        self.button_cpt_val.setEnabled(False)
        self.geol_layers.setEnabled(False)
        self.button_copy_actual.setEnabled(False)
        self.avg_vals.setEnabled(False)
        self.button_copy_avg.setEnabled(False)
        self.full_export.setEnabled(False)


    def enable_buttons(self):
        self.button_depth.setEnabled(True)
        self.cpt_table.setEnabled(True)
        self.button_cpt_val.setEnabled(True)
        self.geol_layers.setEnabled(True)
        self.button_copy_actual.setEnabled(True)
        self.button_copy_avg.setEnabled(True)
        self.avg_vals.setEnabled(True)
        self.full_export.setEnabled(True)


    def eventFilter(self, object: QObject, event: QEvent) -> bool:
        if event.type() == QEvent.Type.WindowStateChange:
            maximized = bool(Qt.WindowState.WindowMaximized & self.windowState())
            self.config['Window']['maximized'] = str(maximized)
            with open('assets/settings.ini', 'w') as configfile: 
                self.config.write(configfile)
        
        return super().eventFilter(object, event)
    
    def resizeEvent(self, event: QResizeEvent) -> None:
        if not self.resizing:
            self.resizing = True
            timer = QTimer()
            timer.singleShot(500,self.on_resize_timer)
            timer.start()
        
        return super().resizeEvent(event)
    
    def on_resize_timer(self):
        if self.isMaximized():
            self.resizing = False
            return
        
        width = str(self.size().width())
        height = str(self.size().height())
        
        self.config['Window']['width'] = width
        self.config['Window']['height'] = height
        with open('assets/settings.ini', 'w') as configfile: 
            self.config.write(configfile)
        self.resizing = False
    
    def set_size(self):
        self.resizing = True
        
        if self.config.getboolean('Window','maximized',fallback=False):
            width = self.config['Window']['width']
            height = self.config['Window']['height']
            self.resize(QSize(int(width),int(height)))
            self.showMaximized()
            return
        
        width = self.config['Window']['width']
        height = self.config['Window']['height']
        self.resize(QSize(int(width),int(height)))
        self.resizing = False
        

def main():
    app = QtWidgets.QApplication([sys.argv])
    QtGui.QFontDatabase.addApplicationFont("assets/fonts/Roboto.ttf")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()